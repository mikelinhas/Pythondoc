from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook("stock.xlsx", read_only=True)
ws = wb["Data Mixto"]

# First ROW --> 5
row0 = 4

# Last ROW
last_row = ws.max_row

# Columns
BAVIERA_CODE_COL = 4 # D
BAVIERA_DESCR_COL = 5 # E
SUPPLIER_COL = 6 # F
COSTE_CANT_COL = 8 # H
PRICE_COL = 43 # AQ


# go through EXCEL to get info
def get_data(code):
	print("Codigo ", code)
	for row in range(row0, last_row):
		baviera_code = ws.cell(row=row, column=BAVIERA_CODE_COL).value
		if code == baviera_code:
			baviera_description = ws.cell(row=row, column=BAVIERA_DESCR_COL).value
			supplier = ws.cell(row=row, column=SUPPLIER_COL).value
			coste_cant = ws.cell(row=row, column=COSTE_CANT_COL).value
			price = ws.cell(row=row, column=PRICE_COL).value
			if coste_cant == "Coste":
				print(code, " - ", baviera_description, " - ", supplier, ": ", price, "€")

get_data("AACCB49050")