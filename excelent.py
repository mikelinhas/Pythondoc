from openpyxl import Workbook

wb = Workbook()

# --- SHEETS ---

# grab the active worksheet
ws = wb.active
ws.title = "New Title" 
ws0 = wb["New Title"]
ws0.sheet_properties.tabColor = "1072BA"

# create new worksheets
ws1 = wb.create_sheet("Sheet1") #insert at end (default)
ws2 = wb.create_sheet("Sheet2", 0) # insert first

print(wb.sheetnames)


# --- CELLS ---

# cells can be accessed directly
cell = ws0["A4"]

# and modified directly
cell.value = 4
ws0["A5"] = 5

#another option
d = ws0.cell(row=4, column=3, value=10)

# ranges of cells can be accessed using slices
cell_range = ws0["A1":"C2"]
colC = ws0["C"]
col_range = ws0["C:D"]
row10 = ws[10]

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
	for cell in row:
		print(cell)

for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
	for cell in col:
		print(cell)

# --- SAVING ---
wb.save("wboo.xlsx")